import urllib
import shutil
from os import makedirs, remove, listdir
from os.path import abspath, exists
from datetime import date
import pandas as pd
import openpyxl
import openpyxl.styles
from sqlalchemy import create_engine
import charts as charts
import fr_charts as fr
import grand_public as gd

ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

# SQL IDPE Connect
engine = create_engine('mysql://labonneboite:%s@127.0.0.1:3306/labonneboite' %
                       urllib.parse.quote_plus('LaB@nneB@ite'))
engine.connect()

# For the evolution of number of IDPEC
query = '''
SELECT count(DISTINCT idutilisateur_peconnect) as count_distinct_idpe, MONTH(dateheure), YEAR(dateheure)
FROM idpe_connect
GROUP BY MONTH(dateheure), YEAR(dateheure)
ORDER BY YEAR(dateheure), MONTH(dateheure);
'''
idpe_connect = pd.read_sql_query(query, engine)
month_idpe = idpe_connect['MONTH(dateheure)'].tolist()
year_idpe = idpe_connect['YEAR(dateheure)'].tolist()

month_year_idpe = []  # Formatting
i=0
while i < len(month_idpe):
    month_year_idpe.append(str(year_idpe[i])+'/'+str(month_idpe[i]))
    i+=1
idpe_connect['Date'] = month_year_idpe

# Count all distinct IDPEC
query_bis = '''
SELECT count(DISTINCT idutilisateur_peconnect)
FROM idpe_connect;
'''
total_idpe_connect = pd.read_sql_query(query_bis, engine)

# Count only significative activity
query_ter = '''
SELECT count(DISTINCT idutilisateur_peconnect)
FROM activity_logs;
'''
total_idpe_connect_sign = pd.read_sql_query(query_ter, engine)


print('SQL query done !')

# Creation of all directory needed
path = abspath('clean_tre.py')[:-12]
try:
    shutil.rmtree(path+'images/')
except:
    pass
try:
    shutil.rmtree(path+'gd_pub/')
except:
    pass
try:
    shutil.rmtree(path+'Clean/')
except:
    pass

makedirs(path+'images/')
makedirs(path+'gd_pub/')
makedirs(path+'Clean/')


# Load CSV and rename columns,values etc.. :
act_dpae = pd.read_csv(path+'act_dpae.csv',
                            sep='|',
                            header=0)

# remove duplicates when multiple activities for the same dpae
act_dpae_bis = act_dpae[act_dpae.premiere_embauche == 'Embauche']
act_dpae_1 = act_dpae_bis.sort_values('dateheure')
act_dpae_2 = act_dpae_1.drop_duplicates(
    subset=['idutilisateur-peconnect', 'siret'], keep='first')

# rename some columns
act_dpae_2.rename(columns={'dateheure': 'date_activite',
                           'kd_dateembauche': 'date_embauche',
                           'nbrjourtravaille': 'duree_activite_cdd_jours',
                           'kn_trancheage': 'tranche_age',
                           'duree_pec': 'duree_prise_en_charge',
                           'dc_commune_id': 'code_postal'
                           },
                  inplace=True)


def get_type_contrat(row):
    if row['dc_typecontrat_id'] == 1:
        return 'CDD'
    elif row['dc_typecontrat_id'] == 2:
        return 'CDI'
    return 'CTT'


act_dpae_2['type_contrat'] = act_dpae_2.apply(
    lambda row: get_type_contrat(row), axis=1)

# TODO : use lambdas functions


def get_nb_mois(row):
    return row['duree_activite_cdd_jours'] // 30


act_dpae_2['duree_activite_cdd_mois'] = act_dpae_2.apply(
    lambda row: get_nb_mois(row), axis=1)


def get_nbr_jours_act_emb(row):
    de = row['date_embauche'][:10].split('-')
    da = row['date_activite'][:10].split('-')
    f_date = date(int(da[0]), int(da[1]), int(da[2]))
    l_date = date(int(de[0]), int(de[1]), int(de[2]))
    delta = l_date - f_date
    return delta.days


act_dpae_2['diff_activite_embauche_jrs'] = act_dpae_2.apply(
    lambda row: get_nbr_jours_act_emb(row), axis=1)


def get_priv_pub(row):
    if row['dc_privepublic'] == 0:
        return 'Public'
    return 'Prive'


act_dpae_2['dc_privepublic'] = act_dpae_2.apply(
    lambda row: get_priv_pub(row), axis=1)


def good_format(row):
    return row['date_embauche'][:-2]


act_dpae_2['date_embauche'] = act_dpae_2.apply(
    lambda row: good_format(row), axis=1)


def del_interrogation(row):
    if row['tranche_age'] == 'de 26 ans ? 50 ans':
        return 'entre 26 et 50 ans'
    return row['tranche_age']


act_dpae_2['tranche_age'] = act_dpae_2.apply(
    lambda row: del_interrogation(row), axis=1)


def del_cdd_incoherent(row):
    try:
        if int(row['duree_activite_cdd_jours']) > 1200:
            return 1
        return 0
    except:
        return 0


act_dpae_2['temporaire'] = act_dpae_2.apply(
    lambda row: del_cdd_incoherent(row), axis=1)
act_dpae_2_bis = act_dpae_2[act_dpae_2.temporaire == 0]

# We only have activities in august for 31/08/2018 --> ugly charts, we want to start from the 1st september
act_dpae_2_bis = act_dpae_2_bis[act_dpae_2_bis.date_activite > "2018-08-31"]

cols_of_interest = ['date_activite',
                    'date_embauche',
                    'type_contrat',
                    'duree_activite_cdd_mois',
                    'duree_activite_cdd_jours',
                    'diff_activite_embauche_jrs',
                    'dc_lblprioritede',
                    'tranche_age',
                    'dc_privepublic',
                    'duree_prise_en_charge',
                    'dn_tailleetablissement',
                    'code_postal']

act_dpae_3 = act_dpae_2_bis[cols_of_interest]


#####
# CHARTS
#####

# Names of different legend for Cohortes
all_the_names_1 = ("Total from activity", "Nbre_Total_DPAE", "Mois d'activité",
                   "Mois d'embauche", "Origine de l'Activité,  en fonction du mois d'Embauche")
all_the_names_2 = ('Nbre_Total_DPAE', "Total from activity", "Mois d'embauche",
                   "Mois d'activité", "Nombres et Mois d'Embauche,  en fonction du mois d'Activité")


def location(num_image, file_name, link=False):  # Pasting of pictures
    if link is True:
        ws.merge_cells("A21:F21")
        ws.merge_cells("G21:L21")
        ws.merge_cells("M21:R21")

        num_image += 3
    list_abscisses = ["A", "G", "M"]
    if "cohorte" in file_name:
        list_abscisses = ["A", "G"]
    y = (((num_image//len(list_abscisses)))*20)+1
    x = list_abscisses[num_image % len(list_abscisses)]
    return x+str(y)

# Write Datas and Charts in Excel :


# Initialisation
sheet_names = [None, "BoxPlot + Graph",
               "Détail Embauches", "Pie Charts", "Map", "Cohortes"]
sheet_sizes = [None, 5, 2, 2, 3]  # Number of files per sheet (start at 0)
num_sheet = 1

# Writes raw data
wb = openpyxl.Workbook()
wb.save('Temporaire.xlsx')
temporaire_df = pd.ExcelWriter(path+'Temporaire.xlsx', engine='xlsxwriter')
act_dpae_3.to_excel(temporaire_df, 'DPAE', index=False)
temporaire_df.save()
book = openpyxl.load_workbook('Temporaire.xlsx', data_only=True)

# Extend columns
for i in range(len(act_dpae_3.columns.tolist())):
    book.active.column_dimensions[ALPHABET[i]].width = 20

# Past of graphics/maps/Pie etc...
book.create_sheet(sheet_names[num_sheet])
ws = book.worksheets[num_sheet]
dict_charts = {('01', "diff_activite_embauche_jrs", "Nombre de Jours entre l'activite sur lbb et la DPAE", "act_emb", 1): charts.BoxPlot,
               ('02', 'duree_activite_cdd_jours', "Durée du CDD obtenu", "cdd_duree", 0): charts.BoxPlot,
               ('03', 'count_distinct_idpe', "Nbre d'IDPE connect par mois", "idpe_connect", 'Date'): charts.graph_sql,
               ('04', 'date_activite', "Nombre d'activités entrainant une DPAE par mois", "act_mo", "month"): charts.Graph,
               ('05', 'date_embauche', "Nombre d'embauches par mois", "emb_mo_gd_public_graph", 'month'): charts.Graph,
               ('07', 'date_embauche', " : Nombre d'embauches par semaine", "emb_sem", 'week'): charts.Graph,
               ('08', "type_contrat", "Type de contrat obtenu", "type_cont_gd_public_pie", None): charts.Pie,
               ('10', "tranche_age", "Pourcentage des differentes tranches d'ages dans les DPAE", "age_gd_public_pie", None): charts.Pie,
               ('11', "dc_privepublic", "Pourcentage d'embauche dans le privé et dans le public", "prive_pub_gd_public_pie", None): charts.Pie,
               ('12', "code_postal", "Part des DPAE par anciennes régions", "old_region_gd_public_svg", "old_region"): fr.map_fr,
               ('13', "code_postal", "Part des DPAE par nouvelles régions", "new_region_gd_public_svg", "new_region"): fr.map_fr,
               ('14', "code_postal", "Part des DPAE par département", "dep_gd_public_svg", "departement"): fr.map_fr,
               ('15', 'date_embauche', all_the_names_1, 'cohorte_1_gd_public', 'date_activite'): charts.Stacked_Bar,
               ('16', 'date_activite', all_the_names_2, 'cohorte_2_gd_public', 'date_embauche'): charts.Stacked_Bar}


####Boucle ? #################################################################################
# Add number of DPAE with LBB activity
nbre_DPAE = act_dpae_3['date_activite'].describe().tolist()[0]
ws.merge_cells('A4:F4')
ws.merge_cells('A5:F5')
cell_A4 = ws['A4']
cell_A4.value = 'Nbre de DPAE ayant pour origine une activité sur LBB'
cell_A4.font = openpyxl.styles.Font(size=10, bold=True, underline='double')
cell_A4.alignment = openpyxl.styles.Alignment(horizontal="center")
cell_A5 = ws['A5']
cell_A5.value = nbre_DPAE
cell_A5.font = openpyxl.styles.Font(size=10, italic=True)
cell_A5.alignment = openpyxl.styles.Alignment(horizontal="center")
DPAE_for_gd_pub = [cell_A4.value, nbre_DPAE]  # for grand_public

# Add number of IDPE unique
nbre_DPAE = total_idpe_connect.loc[0][0]
ws.merge_cells('A7:F7')
ws.merge_cells('A8:F8')
cell_A7 = ws['A7']
cell_A7.value = "Nbre d'IDPE connect unique sur LBB depuis 09/18"
cell_A7.font = openpyxl.styles.Font(size=10, bold=True, underline='double')
cell_A7.alignment = openpyxl.styles.Alignment(horizontal="center")
cell_A8 = ws['A8']
cell_A8.value = nbre_DPAE
cell_A8.font = openpyxl.styles.Font(size=10, italic=True)
cell_A8.alignment = openpyxl.styles.Alignment(horizontal="center")
IDPE_for_gd_pub = [cell_A7.value, nbre_DPAE]  # for grand_public

# Add number of IDPE unique with significative activity
nbre_IDPE_sign = total_idpe_connect_sign.loc[0][0]
ws.merge_cells('A10:F10')
ws.merge_cells('A11:F11')
ws.merge_cells('A12:F12')
cell_A10 = ws['A10']
cell_A10.value = "Nbre d'IDPE connect unique sur LBB, ayant cliqué sur:"
cell_A10.font = openpyxl.styles.Font(size=10, bold=True, underline='double')
cell_A10.alignment = openpyxl.styles.Alignment(horizontal="center")
cell_A11 = ws['A11']
cell_A11.value = "'Favoris','Telecharger PDF','Details d'une entreprise' "
cell_A11.font = openpyxl.styles.Font(size=10, bold=True, underline='double')
cell_A11.alignment = openpyxl.styles.Alignment(horizontal="center")
cell_A12 = ws['A12']
cell_A12.value = nbre_IDPE_sign
cell_A12.font = openpyxl.styles.Font(size=10, italic=True)
cell_A12.alignment = openpyxl.styles.Alignment(horizontal="center")
IDPE_sign_for_gd_pub = [cell_A10.value,
                        nbre_IDPE_sign, cell_A11.value]  # for grand_public

##################################################################################################

num_im = 1
package_svg = []
all_stats = []

for args in dict_charts:  # Creation and saving of charts, using charts.py
    if 'sql' in dict_charts[args].__name__:  # choose data
        data = idpe_connect
    else:
        data = act_dpae_3

    # Creation of charts/maps in directory "images/"
    image = dict_charts[args](
        args[0], args[1], data, args[2], args[3], args[4])  # function

    # if sem in args, function return the number of graph by "week". It means that create a new sheet is necessary
    # if sem not in args AND image is not None, it means that the function return a list of stats (for gd_pub sheet)
    if "sem" in args[3]:
        sheet_sizes.insert(2, image-1)
    elif image is not None:
        all_stats.append(image)

# Iterate through the created images
# Pasting of charts from the directory
for filename in sorted(listdir(path+'images/')):

    img = openpyxl.drawing.image.Image(path+'images/'+filename)

    if "gd_public" in filename:
        shutil.copyfile(path+'images/'+filename, path+'gd_pub/'+filename)

    if "table" in filename:  # it's the table of cohorte --> it's a different size
        img.anchor = 'H1'
        img.height = 750
        img.width = 900
    else:
        # using the function location in order to place the charts
        img.anchor = location(num_im, filename)
        img.height = 400
        img.width = 500

    ws.add_image(img)  # Pasting

    # if it's map --> pasting web link below charts
    if exists(path+filename[:-3]+'svg'):
        cells_link = ws[location(num_im, filename, True)]
        cells_link.hyperlink = filename[:-3]+'svg'
        cells_link.font = openpyxl.styles.Font(
            size=5.5, italic=True, underline='single')
        cells_link.alignment = openpyxl.styles.Alignment(horizontal="center")
        package_svg.append((path, filename[:-3]+'svg'))

    num_im += 1

    # if it's the last charts of the sheet --> change sheet
    if num_im == (sheet_sizes[num_sheet]+1):
        try:
            num_sheet += 1
            book.create_sheet(sheet_names[num_sheet])
            ws = book.worksheets[num_sheet]
            num_im = 0
        except:
            pass

book.save('Impact_lbb_DPAE.xlsx')

# gd_pub sheet
gd.build_grand_public_sheet(DPAE_for_gd_pub,
                            IDPE_for_gd_pub,
                            IDPE_sign_for_gd_pub,
                            all_stats,
                            book,
                            path+'gd_pub/')

# Remove all files/directory useless and create "Clean" package
shutil.rmtree(path+'images/')
shutil.rmtree(path+'gd_pub/')
remove("Temporaire.xlsx")
shutil.copyfile(path+'Impact_lbb_DPAE.xlsx', path +
                'Clean/'+'Impact_lbb_DPAE.xlsx')
for path, svg in package_svg:
    shutil.copyfile(path+svg, path+'Clean/'+svg)
remove("table.html")
for last_files in listdir(path):
    try:
        extension = last_files[last_files.index('.'):]
        if extension == '.svg' or extension == '.xlsx':
            remove(last_files)
    except:
        pass  # It's a directory

def run_main():
    return 0