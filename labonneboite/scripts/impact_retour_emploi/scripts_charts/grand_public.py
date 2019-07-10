import os
import openpyxl


stat_name_in_sheet = [
    'Mediane',
    'Dualité Moyenne/Médiane',
    'Quartile',
    'Ecart-type']
explication = ['La Mediane est le nombre tel que la moitié des valeurs lui sont inférieures',
               'La Moyenne étant souvent biaisée par des valeurs extrêmes, la Médiane est souvent plus représentative',
               'Un Quartile est le nombre tel que 1/4 des valeurs lui sont inférieures',
               "L'Ecart-type permet de mesurer la dispersion d'une série statistique par rapport à la Moyenne"
               ]

name_stats_in_list = [None, "Population", "Moyenne",
                      "Ecart-Type", 'Min', 'Q1', 'Mediane', 'Q3', 'Max']
name_stats_useful = ["Moyenne", "Ecart-Type", 'Min', 'Mediane', 'Max']
name_in_sheet = ["Moyenne :", "La plupart des valeurs se situent dans l'intervalle",
                 "La plus petite des valeurs est :", "La moitié des valeurs sont inférieures à :",
                 "La plus grande valeur est : "]


def build_grand_public_sheet(nbre_DPAE, nbre_IDPE, nbre_IDPE_sign, all_stats, impact_xlsx, path):
    impact_xlsx.create_sheet("Grand Public")
    ws = impact_xlsx.worksheets[-1]

    # Point Stat -------------------------------------

    # Title
    ws.merge_cells('A1:K2')
    cell_A1 = ws['A1']
    cell_A1.value = "Point Statistique : "
    cell_A1.font = openpyxl.styles.Font(size=12, bold=True, underline='single')
    cell_A1.alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center")
    i = 0
    while i < len(explication):
        # Name
        ws.merge_cells('A{}:C{}'.format(i+3, i+3))
        cell_Ax = ws['A{}'.format(i+3)]
        cell_Ax.value = stat_name_in_sheet[i]
        cell_Ax.font = openpyxl.styles.Font(size=10, bold=True)
        cell_Ax.alignment = openpyxl.styles.Alignment(horizontal="center")
        # Définititon
        ws.merge_cells('D{}:K{}'.format(i+3, i+3))
        cell_Dx = ws['D{}'.format(i+3)]
        cell_Dx.value = explication[i]
        cell_Dx.font = openpyxl.styles.Font(size=8, italic=True)

        i += 1

    # DPAE/IDPEC/IDPEC significative

    cells_from_other_sheet = [nbre_DPAE, nbre_IDPE, nbre_IDPE_sign]
    index_abs = 0
    abscisses = [1, 2, 4, 5, 7, 9, 8]

    # Write on the top the count of DPAE/IDPEC etc...
    for vecteur_indicator in cells_from_other_sheet:
        index_vect = 0
        while index_vect < len(vecteur_indicator):
            ws.merge_cells('M{}:R{}'.format(
                abscisses[index_abs], abscisses[index_abs]))
            cell_Mx = ws['M{}'.format(abscisses[index_abs])]
            cell_Mx.value = vecteur_indicator[index_vect]
            if index_vect % 2 != 0:  # if it's the value of the indicator
                cell_Mx.font = openpyxl.styles.Font(size=10, italic=True)
            else:  # if it's the name of the indicator
                cell_Mx.font = openpyxl.styles.Font(
                    size=10, bold=True, underline='double')
            cell_Mx.alignment = openpyxl.styles.Alignment(horizontal="center")
            index_abs += 1
            index_vect += 1

    # Stats BoxPlot -------------------------------------
    y = 10
    for stats_from_boxplot in all_stats:

        # Création of interval [moyenne-ecart_type;moyenne+ecart_type]
        borne_inf = stats_from_boxplot[2]-stats_from_boxplot[3]
        if borne_inf < 0:
            borne_inf = 0
        borne_sup = stats_from_boxplot[2]+stats_from_boxplot[3]
        stats_from_boxplot[3] = "[{}:{}]".format(borne_inf, borne_sup)

        # Writing origin of stats in title
        ws.merge_cells('L{}:R{}'.format(y, y))
        cell_title = ws['L{}'.format(y)]
        cell_title.value = stats_from_boxplot[0]
        cell_title.font = openpyxl.styles.Font(
            size=8, bold=True, underline='double')
        cell_title.alignment = openpyxl.styles.Alignment(horizontal="center")
        y += 2

        count = 0
        for j in range(1, len(stats_from_boxplot)):  # writing all the stats...
            # ...Only if we think that it's intersting (forget the quartile)
            if name_stats_in_list[j] in name_stats_useful:
                # On the right we write the definition
                ws.merge_cells('M{}:P{}'.format(y, y))
                cell_explication = ws['M{}'.format(y)]
                cell_explication.value = name_in_sheet[count]
                cell_explication.font = openpyxl.styles.Font(
                    size=7, bold=None, underline='single')
                cell_explication.alignment = openpyxl.styles.Alignment(
                    horizontal="right")
                # On the left, we write the number corresponding
                ws.merge_cells('Q{}:R{}'.format(y, y))
                cell_number = ws['Q{}'.format(y)]
                try:
                    cell_number.value = round(stats_from_boxplot[j], 2)
                except:  # Formatting of the interval
                    cell_number.value = "["+str(round(float(stats_from_boxplot[j][1:stats_from_boxplot[j].index(":")]), 2))+" : "+str(
                        round(float(stats_from_boxplot[j][stats_from_boxplot[j].index(":")+1:-1]), 2))+"]"
                cell_number.font = openpyxl.styles.Font(size=9, italic=True)
                cell_number.alignment = openpyxl.styles.Alignment(
                    horizontal='right')
                y += 1
                count += 1

        y += 1

    # Charts -------------------------------------------------------
    dico_files_types = {}
    dico_types_location = {'png': ['A8', 'E8', 'I8'], 'svg': [
        'A24', 'E24', 'I24'], 'graph': ['N26', 'N40'], 'cohorte': ['B41', 'H41']}
    index = {'png': 0, 'svg': 0, 'graph': 0, 'cohorte': 0}
    cells_merge_for_link = ['A38:D38', 'E38:H38', 'I38:L38']
    link = ['A20', 'G20', 'M20']
    index_link = 0
    taille = (250, 300)

    # corresponding all files of the directory with their types
    for files in os.listdir(path):
        if 'table' not in files:
            # we need to add the cohorte's table at the directory, but we don't want to show it in gd_pub
            if 'graph' in files:
                dico_files_types[files] = 'graph'
            elif 'svg' in files:
                dico_files_types[files] = 'svg'
            elif 'cohorte' in files:
                dico_files_types[files] = 'cohorte'
            else:
                dico_files_types[files] = files[-3:]

    for files_pasted in dico_files_types:  # Pasting on the sheet
        img = openpyxl.drawing.image.Image(path+files_pasted)
        types = dico_files_types[files_pasted]
        img.anchor = dico_types_location[types][index[types]]
        img.height = taille[0]
        img.width = taille[1]
        ws.add_image(img)

        if types == 'svg':  # if svg, pasting the web link below
            ws.merge_cells(cells_merge_for_link[index_link])
            cell_link = ws[cells_merge_for_link[index_link][:3]]
            cell_link.hyperlink = impact_xlsx.worksheets[4][link[index_link]].value
            cell_link.font = openpyxl.styles.Font(
                size=5.5, italic=True, underline='single')
            cell_link.alignment = openpyxl.styles.Alignment(
                horizontal="center")
            index_link += 1

        index[types] += 1

    impact_xlsx.save('Impact_lbb_DPAE.xlsx')
